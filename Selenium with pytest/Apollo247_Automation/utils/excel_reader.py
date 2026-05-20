from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def read_excel(file_path, sheet_name="Sheet1"):
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        data = []

        # skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data